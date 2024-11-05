from flask import Flask, render_template, request, redirect, url_for, session, flash, get_flashed_messages
import openpyxl
import requests
import locale
from datetime import datetime
from dotenv import load_dotenv
import os 

load_dotenv()  # Carrega as variáveis do arquivo .env

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

# Configuração da localidade para o formato de moeda brasileiro
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

# Credenciais de login
USUARIO = os.getenv('USUARIO')
SENHA = os.getenv('SENHA')

MESES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

def carregar_planilha(ano):
    try:
        workbook = openpyxl.load_workbook(f"investimentos_btc_{ano}.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        for mes in MESES:
            sheet = workbook.create_sheet(title=mes)
            sheet.append(["Investimento (R$)", "Quantidade de BTC"])
        workbook.save(f"investimentos_btc_{ano}.xlsx")
    return workbook

def obter_cotacao_btc():
    try:
        url = "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=brl"
        response = requests.get(url).json()
        return response['bitcoin']['brl']
    except:
        return None

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == USUARIO and password == SENHA:
            session['logged_in'] = True
            flash("Login bem-sucedido! Redirecionando...")
            return redirect(url_for('loading'))
        else:
            flash('Credenciais inválidas. Tente novamente.')
    return render_template('login.html')

@app.route('/loading')
def loading():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template("loading.html")

@app.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template("index.html")

@app.route('/selecionar_ano', methods=['GET', 'POST'])
def selecionar_ano():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        ano = request.form['ano']
        return redirect(url_for('resumo', ano=ano))
    anos = [str(datetime.now().year - i) for i in range(5)]
    return render_template("selecionar_ano.html", anos=anos)

@app.route('/resumo/<ano>')
def resumo(ano):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    workbook = carregar_planilha(ano)
    resumo_mensal = []

    for sheet in workbook.sheetnames:
        aba = workbook[sheet]
        investimento = sum(cell.value for cell in aba["A"] if isinstance(cell.value, (int, float)))
        quantidade_btc = sum(cell.value for cell in aba["B"] if isinstance(cell.value, (int, float)))
        investimento_formatado = locale.currency(investimento, grouping=True)
        resumo_mensal.append({"mes": sheet, "investimento": investimento_formatado, "quantidade_btc": quantidade_btc})

    return render_template("resumo.html", resumo_mensal=resumo_mensal, ano=ano)

@app.route('/lucro')
def lucro():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    anos = [str(datetime.now().year - i) for i in range(5)]
    total_investido = 0
    total_btc = 0

    for ano in anos:
        workbook = carregar_planilha(ano)
        for sheet in workbook.sheetnames:
            aba = workbook[sheet]
            total_investido += sum(cell.value for cell in aba["A"] if isinstance(cell.value, (int, float)))
            total_btc += sum(cell.value for cell in aba["B"] if isinstance(cell.value, (int, float)))

    cotacao_btc = obter_cotacao_btc()
    if cotacao_btc is None:
        flash("Erro ao obter a cotação do BTC.")
        lucro_prejuizo = None
    else:
        valor_total_em_reais = total_btc * cotacao_btc
        lucro_prejuizo = valor_total_em_reais - total_investido

        total_investido = locale.currency(total_investido, grouping=True)
        valor_total_em_reais = locale.currency(valor_total_em_reais, grouping=True)
        lucro_prejuizo_formatado = locale.currency(lucro_prejuizo, grouping=True)
        cotacao_btc = locale.currency(cotacao_btc, grouping=True)

    return render_template(
        "lucro.html",
        total_investido=total_investido,
        total_btc=total_btc,
        valor_total_em_reais=valor_total_em_reais if cotacao_btc else None,
        lucro_prejuizo_formatado=lucro_prejuizo_formatado if cotacao_btc else None,
        cotacao_btc=cotacao_btc,
        lucro_prejuizo=lucro_prejuizo
    )

@app.route('/editar', methods=['GET', 'POST'])
def editar():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    ano = request.args.get('ano')
    if request.method == 'POST':
        mes_selecionado = request.form.get('mes')
        workbook = carregar_planilha(ano)
        aba = workbook[mes_selecionado]
        valores_atuais = [{"investimento": row[0], "quantidade_btc": row[1]} for row in aba.iter_rows(min_row=2, values_only=True)]
        return render_template("editar_mes.html", mes_selecionado=mes_selecionado, valores_atuais=valores_atuais, ano=ano)
    
    return render_template("selecionar_mes.html", meses=MESES, ano=ano)

@app.route('/salvar_edicao', methods=['POST'])
def salvar_edicao():
    ano = request.form.get('ano')
    mes_selecionado = request.form.get('mes')
    novos_investimentos = request.form.getlist('investimento[]')
    novas_quantidades_btc = request.form.getlist('quantidade_btc[]')

    workbook = carregar_planilha(ano)
    aba = workbook[mes_selecionado]

    aba.delete_rows(2, aba.max_row)

    for investimento, quantidade_btc in zip(novos_investimentos, novas_quantidades_btc):
        if investimento and quantidade_btc:
            investimento = float(investimento.replace(',', '.'))
            quantidade_btc = float(quantidade_btc.replace(',', '.'))
            aba.append([investimento, quantidade_btc])

    workbook.save(f"investimentos_btc_{ano}.xlsx")
    flash("Alterações salvas com sucesso!")
    return redirect(url_for('resumo', ano=ano))

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    get_flashed_messages()
    flash("Você foi desconectado.")
    return redirect(url_for('login'))

if __name__ == "__main__":
        app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))

